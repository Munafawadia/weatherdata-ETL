# -*- coding: utf-8 -*-
"""
Created on Sun Sep 12 21:42:44 2021

@author: awadi
"""

from tkinter import *
from tkinter import ttk
import pandas as pd
import tkinter.font as font
import io
import requests
import openpyxl
from openpyxl import load_workbook
from tkinter import filedialog
from tkcalendar import Calendar,DateEntry

window = Tk()
window.geometry("500x500")
#window.iconphoto(False, icon)
window.title("Gas Tracker App")
window.configure()
window.iconbitmap("icon.ico")   

window.grid_rowconfigure((0,1), weight = 1)
window.grid_columnconfigure(0, weight = 1)

font1 = ("Calibiri", 12)

#Weather Frame

weather_frame = LabelFrame(window, text = " Step 1 - Weather Data ", font = font1, fg = "black")
weather_frame.grid(row =0, column = 0, sticky = "nsew", padx = 10, pady = 10)
weather_frame.grid_columnconfigure((0,1,2), weight = 1)
weather_frame.grid_rowconfigure((0,1), weight = 1)

file_label = Label(weather_frame, text = "File", fg = "black")
file_label.grid(row = 0, column = 0)

global filepath_var
filepath_var = StringVar()
file_path_entry = Entry(weather_frame, width = 20, bg = "grey85", textvariable = filepath_var, fg = "black")
file_path_entry.grid(row = 0, column = 1)

def browse():
    global file_path
    file_path = filedialog.askopenfilename(filetypes=(("excel","*.xlsx"),("All files","*.*")))
    filepath_var.set(file_path)
    
browse = Button(weather_frame, text = "Browse file", command = browse,  fg = "black")
browse.grid(row = 0, column = 2)

select_time_label = Label(weather_frame, text = "Select Month/Year",  fg = "black")
select_time_label.grid(row = 1, column = 0)

month_variable = StringVar(weather_frame)
monthoptions = ["Select Month", "01", "02", "03", "04", "05", "06", "07", "08", "09", "10", "11", "12"]
month_menu = ttk.OptionMenu(weather_frame, month_variable, *monthoptions)
month_menu.grid(row = 1, column = 1)

year_var = StringVar(weather_frame)
yearoption = ["Select Year", "2020", "2021", "2022", "2023", "2024", "2025"]
year_menu = ttk.OptionMenu(weather_frame, year_var, *yearoption)
year_menu.grid(row = 1, column = 2)

def update():
    
    month = month_variable.get()
    year = year_var.get()
    df = pd.read_excel(file_path, "Weather Input 2021")
    url = "https://dd.weather.gc.ca/climate/observations/daily/csv/ON/climate_daily_ON_6158355_"+year+"-"+month+"_P1D.csv"
    s = requests.get(url).content
    c = pd.read_csv(io.StringIO(s.decode('cp1252')))
    c["Date/Time"] = pd.to_datetime(c["Date/Time"])
    
    new_df = pd.concat([df,c]).drop_duplicates().reset_index(drop = True)
    new_df.sort_values(by = "Date/Time", inplace = True)
    
    book = load_workbook(file_path)
    writer = pd.ExcelWriter(file_path, engine = "openpyxl")
    writer.book = book
    std=book.get_sheet_by_name("Weather Input 2021")
    book.remove_sheet(std)
    new_df.to_excel(writer, "Weather Input 2021", index = False)
    writer.save()
    
update = Button(weather_frame, text = "Update Weather Data in tool", command = update,  fg = "black")
update.grid(row = 2, column = 0, columnspan = 2, pady = 10, sticky = "e")  

consumption_frame = LabelFrame(window, text = " Step 2 - Natural Gas Consumption ", font = font1,  fg = "black")
consumption_frame.grid(row =1, column = 0, sticky = "nsew", padx = 10, pady = 10)
consumption_frame.grid_columnconfigure((0,1), weight = 1)
consumption_frame.grid_rowconfigure((0,1,2,3,4), weight = 1)

class MyDateEntry(DateEntry):
    def drop_down(self):
        """Display or withdraw the drop-down calendar depending on its current state."""
        if self._calendar.winfo_ismapped():
            self._top_cal.withdraw()
        else:
            self._validate_date()
            date = self.parse_date(self.get())
            x = self.winfo_rootx()
            y = self.winfo_rooty() + self.winfo_height()
            if self.winfo_toplevel().attributes('-topmost'):
                self._top_cal.attributes('-topmost', True)
            else:
                self._top_cal.attributes('-topmost', False)
            # - patch begin: make sure the drop-down calendar is visible
            if x+self._top_cal.winfo_width() > self.winfo_screenwidth():
                x = self.winfo_screenwidth() - self._top_cal.winfo_width()
            if y+self._top_cal.winfo_height() > self.winfo_screenheight()-30:
                y = self.winfo_rooty() - self._top_cal.winfo_height()
            # - patch end
            self._top_cal.geometry('+%i+%i' % (x, y))
            self._top_cal.deiconify()
            self._calendar.focus_set()
            self._calendar.selection_set(date)

from_label  = Label(consumption_frame, text = "Begin",  fg = "black")
from_label.grid(row = 0, column =0)

begin_date = MyDateEntry(consumption_frame,width=10, date_pattern = "y-mm-dd",  fg = "black")
begin_date.grid(row = 0, column = 1)

to_label = Label(consumption_frame, text = "End",  fg = "black")
to_label.grid(row = 1, column = 0)

end_date = MyDateEntry(consumption_frame,width=10, date_pattern = "y-mm-dd",  fg = "black")
end_date.grid(row = 1, column = 1)

cons_label = Label(consumption_frame, text = "Natural Gas Consumption", fg = "black")
cons_label.grid(row = 2, column = 0)

cons_entry = Entry(consumption_frame, width = 15, bg = "grey85")
cons_entry.grid(row =2, column = 1)


    
def cons_update():
    
    consumption_df = pd.read_excel(file_path, "data")
    consumption_df.loc[-1] = [begin_date.get(), end_date.get(), float(cons_entry.get())]
    consumption_df["Begin"] = pd.to_datetime(consumption_df["Begin"])
    consumption_df["End"] = pd.to_datetime(consumption_df["End"])
    consumption_df = consumption_df.drop_duplicates().reset_index(drop = True)
       
    book = load_workbook(file_path)
    writer = pd.ExcelWriter(file_path, engine = "openpyxl")
    writer.book = book
    std=book.get_sheet_by_name("data")
    book.remove_sheet(std)
    consumption_df.to_excel(writer, "data", index = False)
    writer.save()

cons_data_bt = Button(consumption_frame, text = "Update Consumption Data", command = cons_update, fg = "black")
cons_data_bt.grid(row = 4, column = 0, columnspan = 2)

message = Label(window, fg = "black")
message.grid(row = 4, column = 0, columnspan = 2)

window.mainloop()